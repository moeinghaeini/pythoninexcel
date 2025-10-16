"""
Script to create the Job Market Explorer Excel workbook
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
import os

def create_excel_workbook():
    """Create the Job Market Explorer Excel workbook"""
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    instruction_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    python_cell_fill = PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Data Input
    ws1 = wb.create_sheet("Data Input")
    ws1['A1'] = "Job Market Explorer - Data Input"
    ws1['A1'].font = Font(bold=True, size=16)
    ws1['A1'].fill = header_fill
    
    ws1['A3'] = "Instructions:"
    ws1['A3'].font = Font(bold=True)
    ws1['A3'].fill = instruction_fill
    
    ws1['A4'] = "1. Load your job data using the Python cell below"
    ws1['A4'].fill = instruction_fill
    ws1['A5'] = "2. Use the sample data from sample_data/jobs_sample.csv"
    ws1['A5'].fill = instruction_fill
    ws1['A6'] = "3. Or load your own CSV/Excel file with job data"
    ws1['A6'].fill = instruction_fill
    
    ws1['A8'] = "Python Code to Load Data:"
    ws1['A8'].font = Font(bold=True)
    ws1['A8'].fill = python_cell_fill
    
    # Add Python code for loading data
    python_code = """# Load job data
import pandas as pd
from python_functions import load_job_data

# Load the sample data
df = load_job_data('sample_data/jobs_sample.csv')
df.head(10)"""
    
    ws1['A9'] = python_code
    ws1['A9'].fill = python_cell_fill
    ws1['A9'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # Sheet 2: Filter Controls
    ws2 = wb.create_sheet("Filter Controls")
    ws2['A1'] = "Job Market Explorer - Filter Controls"
    ws2['A1'].font = Font(bold=True, size=16)
    ws2['A1'].fill = header_fill
    
    ws2['A3'] = "Filter Parameters:"
    ws2['A3'].font = Font(bold=True)
    
    # Filter parameter labels and default values
    filter_params = [
        ("B2", "Job Title Filter", "All"),
        ("B3", "Location Filter", "All"),
        ("B4", "Min Salary", "50000"),
        ("B5", "Max Salary", "150000"),
        ("B6", "Min Experience (years)", "1"),
        ("B7", "Max Experience (years)", "10"),
        ("B8", "Keyword Search", "python")
    ]
    
    for i, (cell, label, default_value) in enumerate(filter_params, start=4):
        ws2[f'A{i}'] = label
        ws2[f'A{i}'].font = Font(bold=True)
        ws2[f'B{i}'] = default_value
        ws2[f'B{i}'].border = border
    
    ws2['A12'] = "Python Code to Apply Filters:"
    ws2['A12'].font = Font(bold=True)
    ws2['A12'].fill = python_cell_fill
    
    filter_code = """# Apply filters based on control values
from python_functions import filter_jobs

# Get filter values from cells
filtered_df = filter_jobs(
    df,
    job_title=xl("B2"),  # Reference to Job Title filter
    location=xl("B3"),   # Reference to Location filter
    min_salary=xl("B4"), # Reference to Min Salary
    max_salary=xl("B5"), # Reference to Max Salary
    min_experience=xl("B6"), # Reference to Min Experience
    max_experience=xl("B7"), # Reference to Max Experience
    keyword=xl("B8")     # Reference to Keyword Search
)

filtered_df"""
    
    ws2['A13'] = filter_code
    ws2['A13'].fill = python_cell_fill
    ws2['A13'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # Sheet 3: Analysis Dashboard
    ws3 = wb.create_sheet("Analysis Dashboard")
    ws3['A1'] = "Job Market Explorer - Analysis Dashboard"
    ws3['A1'].font = Font(bold=True, size=16)
    ws3['A1'].fill = header_fill
    
    # Dashboard sections
    dashboard_sections = [
        ("A3", "Job Summary", "D2", "get_job_summary(filtered_df)"),
        ("A6", "Salary Statistics", "D5", "calculate_salary_stats(filtered_df)"),
        ("A9", "Top Job Titles", "D8", "filtered_df['Job Title'].value_counts().head(5)"),
        ("A13", "Top Locations", "D12", "filtered_df['Location'].value_counts().head(5)"),
        ("A17", "Top Companies", "D16", "filtered_df['Company'].value_counts().head(5)")
    ]
    
    for i, (label_cell, title, code_cell, code) in enumerate(dashboard_sections):
        ws3[label_cell] = title
        ws3[label_cell].font = Font(bold=True)
        ws3[label_cell].fill = instruction_fill
        
        ws3[f'{label_cell[0]}{int(label_cell[1:])+1}'] = f"Python Code in {code_cell}:"
        ws3[f'{label_cell[0]}{int(label_cell[1:])+1}'].font = Font(bold=True)
        ws3[f'{label_cell[0]}{int(label_cell[1:])+1}'].fill = python_cell_fill
        
        ws3[f'{label_cell[0]}{int(label_cell[1:])+2}'] = f"from python_functions import get_job_summary, calculate_salary_stats\n\n{code}"
        ws3[f'{label_cell[0]}{int(label_cell[1:])+2}'].fill = python_cell_fill
        ws3[f'{label_cell[0]}{int(label_cell[1:])+2}'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # Sheet 4: Visualizations
    ws4 = wb.create_sheet("Visualizations")
    ws4['A1'] = "Job Market Explorer - Visualizations"
    ws4['A1'].font = Font(bold=True, size=16)
    ws4['A1'].fill = header_fill
    
    # Visualization sections
    viz_sections = [
        ("A3", "Salary Analysis Chart", "E2", "create_salary_chart(filtered_df)"),
        ("A16", "Location Distribution Chart", "E15", "create_location_chart(filtered_df)"),
        ("A29", "Experience Level Chart", "E28", "create_experience_chart(filtered_df)")
    ]
    
    for i, (label_cell, title, code_cell, code) in enumerate(viz_sections):
        ws4[label_cell] = title
        ws4[label_cell].font = Font(bold=True)
        ws4[label_cell].fill = instruction_fill
        
        ws4[f'{label_cell[0]}{int(label_cell[1:])+1}'] = f"Python Code in {code_cell}:"
        ws4[f'{label_cell[0]}{int(label_cell[1:])+1}'].font = Font(bold=True)
        ws4[f'{label_cell[0]}{int(label_cell[1:])+1}'].fill = python_cell_fill
        
        ws4[f'{label_cell[0]}{int(label_cell[1:])+2}'] = f"from python_functions import create_salary_chart, create_location_chart, create_experience_chart\nimport matplotlib.pyplot as plt\n\nfig = {code}\nplt.show()"
        ws4[f'{label_cell[0]}{int(label_cell[1:])+2}'].fill = python_cell_fill
        ws4[f'{label_cell[0]}{int(label_cell[1:])+2}'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # Sheet 5: Export
    ws5 = wb.create_sheet("Export")
    ws5['A1'] = "Job Market Explorer - Export Results"
    ws5['A1'].font = Font(bold=True, size=16)
    ws5['A1'].fill = header_fill
    
    # Export sections
    export_sections = [
        ("A3", "Export Filtered Data", "F2", "export_filtered_data(filtered_df, 'filtered_jobs_export.xlsx')"),
        ("A6", "Export Summary Report", "F5", "Create and export comprehensive summary report")
    ]
    
    for i, (label_cell, title, code_cell, code) in enumerate(export_sections):
        ws5[label_cell] = title
        ws5[label_cell].font = Font(bold=True)
        ws5[label_cell].fill = instruction_fill
        
        ws5[f'{label_cell[0]}{int(label_cell[1:])+1}'] = f"Python Code in {code_cell}:"
        ws5[f'{label_cell[0]}{int(label_cell[1:])+1}'].font = Font(bold=True)
        ws5[f'{label_cell[0]}{int(label_cell[1:])+1}'].fill = python_cell_fill
        
        if i == 0:  # Export filtered data
            export_code = """from python_functions import export_filtered_data

result = export_filtered_data(filtered_df, 'filtered_jobs_export.xlsx')
result"""
        else:  # Export summary report
            export_code = """import pandas as pd

# Create summary report
report_data = {
    'Metric': ['Total Jobs', 'Average Salary', 'Top Location', 'Top Company'],
    'Value': [
        len(filtered_df),
        f"${filtered_df['Salary'].mean():,.2f}",
        filtered_df['Location'].mode().iloc[0] if not filtered_df.empty else 'N/A',
        filtered_df['Company'].mode().iloc[0] if not filtered_df.empty else 'N/A'
    ]
}

report_df = pd.DataFrame(report_data)
report_df.to_excel('job_market_summary.xlsx', index=False)
"Summary report exported to job_market_summary.xlsx" """
        
        ws5[f'{label_cell[0]}{int(label_cell[1:])+2}'] = export_code
        ws5[f'{label_cell[0]}{int(label_cell[1:])+2}'].fill = python_cell_fill
        ws5[f'{label_cell[0]}{int(label_cell[1:])+2}'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # Add setup instructions
    ws5['A10'] = "Setup Instructions:"
    ws5['A10'].font = Font(bold=True, size=14)
    ws5['A10'].fill = instruction_fill
    
    instructions = [
        "1. Enable Python in Excel: Insert > Get Add-ins > Search 'Python in Excel'",
        "2. Install dependencies: pip install -r requirements.txt",
        "3. Copy Python code from cells above into Python-enabled cells in Excel",
        "4. Load sample data from sample_data/jobs_sample.csv",
        "5. Adjust filter parameters in Sheet 2 to explore the data"
    ]
    
    for i, instruction in enumerate(instructions, start=11):
        ws5[f'A{i}'] = instruction
        ws5[f'A{i}'].fill = instruction_fill
    
    # Adjust column widths
    for ws in [ws1, ws2, ws3, ws4, ws5]:
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 30
    
    return wb

if __name__ == "__main__":
    # Create the workbook
    wb = create_excel_workbook()
    
    # Save the workbook
    output_file = "job_market_explorer.xlsx"
    wb.save(output_file)
    print(f"Excel workbook created successfully: {output_file}")
    print("\nNext steps:")
    print("1. Open the Excel file in Microsoft Excel")
    print("2. Enable Python in Excel if not already done")
    print("3. Copy the Python code from the cells into Python-enabled cells")
    print("4. Run the Python code to load and analyze job data")
